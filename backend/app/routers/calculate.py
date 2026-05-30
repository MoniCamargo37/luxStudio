import re
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, File, UploadFile
from openpyxl import load_workbook

from ..schemas.models import (
    AdvancedOptimizationRequest,
    BatchCalculationItem,
    BatchCalculationResponse,
    CalculationConfig,
    CalculationResult,
    OptimizationResponse,
)
from ..services.calculator import run_calculation
from ..services.ldt_loader import get_all_ldts
from ..services.ldt_matcher import require_ldt_for_config

router = APIRouter()


@router.post("/calculate", response_model=CalculationResult)
async def calculate(config: CalculationConfig):
    """Run a full CIE 140 / EN 13201 calculation for the given configuration."""
    ldt_id, _ = require_ldt_for_config(config)
    return run_calculation(config, ldt_id)


OPTIMIZATION_OBJECTIVE = "Minimize luminaire power while satisfying all active EN 13201 criteria"
ADVANCED_OPTIMIZATION_OBJECTIVE = "Fit the solution close to the EN 13201 limits while staying compliant"
ADVANCED_OBJECTIVE_LABELS = {
    "technical_limits": "Fit the solution close to the EN 13201 limits while staying compliant",
    "min_power": "Minimize luminaire power while satisfying all active EN 13201 criteria",
    "max_spacing": "Maximize pole spacing while satisfying all active EN 13201 criteria",
}
OPTIMIZATION_MIN_POWER = 1.0
OPTIMIZATION_MAX_POWER = 500.0
OPTIMIZATION_PRECISION = 0.1
SPACING_CANDIDATES = [60.0, 55.0, 50.0, 45.0, 40.0, 35.0, 30.0, 25.0, 20.0, 15.0, 10.0, 5.0]
HEIGHT_CANDIDATES = [4.0, 6.0, 8.0, 9.0, 10.0, 12.0, 14.0, 16.0, 18.0, 20.0]
OPTIMIZATION_FIXED_PARAMETERS = [
    "road_width",
    "sidewalk_left",
    "sidewalk_right",
    "lanes",
    "pavement",
    "lighting_class",
    "maintenance_factor",
    "arrangement",
    "spacing",
    "height",
    "arm_length",
    "pole_offset",
    "tilt",
    "power",
    "manufacturer",
    "model_family",
    "optic_family",
    "cct",
]


def _advanced_objective_label(objective: str) -> str:
    return ADVANCED_OBJECTIVE_LABELS.get(objective, ADVANCED_OPTIMIZATION_OBJECTIVE)


def _failed_criteria(result: CalculationResult) -> str:
    failed = [item for item in result.criteria if not item.passed]
    if not failed:
        return "none"
    return ", ".join(
        f"{item.name}: {item.value:.3g} / required {item.required:.3g}"
        for item in failed
    )


def _with_power(config: CalculationConfig, power: float, ldt_id: str) -> CalculationConfig:
    return CalculationConfig(**{
        **config.model_dump(),
        "power": power,
        "ldt_id": ldt_id,
    })


def _with_updates(config: CalculationConfig, updates: dict, ldt_id: str) -> CalculationConfig:
    return CalculationConfig(**{
        **config.model_dump(),
        **updates,
        "ldt_id": ldt_id,
    })


def _unique_candidates(values: list[float], current: float) -> list[float]:
    rounded = {round(value, 2) for value in values}
    rounded.add(round(current, 2))
    return sorted(rounded, reverse=True)


def _fixed_parameters_for(unlocked: set[str]) -> list[str]:
    return [item for item in OPTIMIZATION_FIXED_PARAMETERS if item not in unlocked]


def _advanced_score(result: CalculationResult, original: CalculationConfig, objective: str) -> tuple[float, float, float]:
    movement = abs(result.config.height - original.height) + abs(result.config.spacing - original.spacing)
    if objective == "min_power":
        return (
            result.config.power,
            result.config.power / max(result.config.spacing, 0.1),
            movement,
        )
    if objective == "max_spacing":
        return (
            -result.config.spacing,
            result.config.power / max(result.config.spacing, 0.1),
            result.config.power,
        )
    technical_score, largest_margin = _technical_limit_score(result)
    return (technical_score, largest_margin, movement)


def _technical_limit_score(result: CalculationResult) -> tuple[float, float]:
    margins = []
    for criterion in result.criteria:
        required = float(criterion.required or 0)
        value = float(criterion.value or 0)
        if required <= 0:
            continue
        if criterion.name.upper().startswith("TI"):
            margin = max(0.0, (required - value) / required)
        else:
            margin = max(0.0, (value - required) / required)
        margins.append(margin)

    if not margins:
        return 0.0, 0.0
    return sum(margin * margin for margin in margins), max(margins)


def _optimize_power_for_config(config: CalculationConfig, ldt_id: str) -> tuple[bool, int, CalculationResult, str]:
    checked = 0
    results_by_power: dict[float, CalculationResult] = {}

    def calculate_power(power: float) -> CalculationResult:
        nonlocal checked
        key = round(max(OPTIMIZATION_MIN_POWER, min(OPTIMIZATION_MAX_POWER, power)), 4)
        if key not in results_by_power:
            results_by_power[key] = run_calculation(_with_power(config, key, ldt_id), ldt_id)
            checked += 1
        return results_by_power[key]

    current_power = max(OPTIMIZATION_MIN_POWER, min(OPTIMIZATION_MAX_POWER, float(config.power)))
    current_result = calculate_power(current_power)
    high: Optional[float] = None
    low: float = OPTIMIZATION_MIN_POWER

    if current_result.compliant:
        high = current_power
        probe = current_power
        while probe > OPTIMIZATION_MIN_POWER:
            next_probe = max(OPTIMIZATION_MIN_POWER, probe / 2.0)
            probe_result = calculate_power(next_probe)
            if probe_result.compliant:
                high = next_probe
                probe = next_probe
            else:
                low = next_probe
                break
    else:
        max_power_result = calculate_power(OPTIMIZATION_MAX_POWER)
        if max_power_result.compliant:
            low = current_power
            high = OPTIMIZATION_MAX_POWER

    if high is None:
        max_result = results_by_power.get(OPTIMIZATION_MAX_POWER) or calculate_power(OPTIMIZATION_MAX_POWER)
        return False, checked, current_result, _failed_criteria(max_result)

    while high - low > OPTIMIZATION_PRECISION / 2.0:
        mid = (low + high) / 2.0
        result = calculate_power(mid)
        if result.compliant:
            high = mid
        else:
            low = mid

    final_power = round(high, 1)
    final_result = calculate_power(final_power)

    # Rounding down to one decimal can sit just below the real threshold.
    while not final_result.compliant and final_power < OPTIMIZATION_MAX_POWER:
        final_power = round(final_power + 0.1, 1)
        final_result = calculate_power(final_power)

    return True, checked, final_result, "none"


def _advanced_unlocked(variables) -> set[str]:
    unlocked = set()
    if variables.power:
        unlocked.add("power")
    if variables.spacing:
        unlocked.add("spacing")
    if variables.height:
        unlocked.add("height")
    if getattr(variables, "optic_family", False):
        unlocked.add("optic_family")
    return unlocked


def _run_advanced_search(
    config: CalculationConfig,
    variables,
    objective: str,
    ldt_id: str,
    objective_label: str,
) -> OptimizationResponse:
    spacing_values = _unique_candidates(SPACING_CANDIDATES, config.spacing) if variables.spacing else [config.spacing]
    height_values = _unique_candidates(HEIGHT_CANDIDATES, config.height) if variables.height else [config.height]
    unlocked = _advanced_unlocked(variables)

    checked = 0
    best_result: Optional[CalculationResult] = None
    best_score: Optional[tuple[float, float, float]] = None
    first_failure = "none"
    last_result: Optional[CalculationResult] = None

    for spacing in spacing_values:
        for height in height_values:
            candidate_config = _with_updates(config, {"spacing": spacing, "height": height}, ldt_id)
            if variables.power:
                feasible, candidate_checked, result, failures = _optimize_power_for_config(candidate_config, ldt_id)
                checked += candidate_checked
            else:
                result = run_calculation(candidate_config, ldt_id)
                checked += 1
                feasible = result.compliant
                failures = _failed_criteria(result)

            last_result = result
            if not feasible:
                if first_failure == "none":
                    first_failure = failures
                continue

            score = _advanced_score(result, config, objective)
            if best_score is None or score < best_score:
                best_score = score
                best_result = result

    if best_result is None:
        fallback = last_result or run_calculation(config, ldt_id)
        if last_result is None:
            checked += 1
        return OptimizationResponse(
            feasible=False,
            message=(
                "No compliant solution was found with the selected advanced variables. "
                f"Failing criteria seen during search: {first_failure}."
            ),
            objective=objective_label,
            fixed_parameters=_fixed_parameters_for(unlocked),
            checked=checked,
            config=config,
            result=fallback,
        )

    return OptimizationResponse(
        feasible=True,
        message=(
            f"Best compliant setup found for selected priority: {best_result.config.power:.1f} W at "
            f"{best_result.config.spacing:.1f} m spacing and {best_result.config.height:.1f} m height."
        ),
        objective=objective_label,
        fixed_parameters=_fixed_parameters_for(unlocked),
        checked=checked,
        config=best_result.config,
        result=best_result,
    )


def _optic_candidates(config: CalculationConfig, requested: Optional[list[str]]) -> list[str]:
    requested_set = set(requested or [])
    candidates = [
        item.get("optic_family")
        for item in get_all_ldts()
        if item.get("manufacturer", "Unknown") == config.manufacturer
        and item.get("model_family", "UNKNOWN") == config.model_family
        and item.get("optic_family")
        and (not requested_set or item.get("optic_family") in requested_set)
    ]
    unique = sorted(set(candidates))
    return unique or [config.optic_family]


@router.post("/optimize/simple", response_model=OptimizationResponse)
async def optimize_simple(config: CalculationConfig):
    """Find the lowest compliant power while keeping the selected geometry and luminaire fixed."""
    ldt_id, _ = require_ldt_for_config(config)

    if ldt_id.startswith("temp-"):
        exact_result = run_calculation(config, ldt_id)
        return OptimizationResponse(
            feasible=False,
            message=(
                "External LDTs are calculated exactly. Auto optimize v1 uses the DB reference "
                "and power/CCT formulas, so select a catalog luminaire to optimize power."
            ),
            objective=OPTIMIZATION_OBJECTIVE,
            fixed_parameters=_fixed_parameters_for({"power"}),
            checked=1,
            config=config,
            result=exact_result,
        )

    feasible, checked, result, failures = _optimize_power_for_config(config, ldt_id)
    if not feasible:
        return OptimizationResponse(
            feasible=False,
            message=(
                "No compliant solution was found by changing power only up to 500 W. "
                f"Failing criteria at the upper limit: {failures}."
            ),
            objective=OPTIMIZATION_OBJECTIVE,
            fixed_parameters=_fixed_parameters_for({"power"}),
            checked=checked,
            config=config,
            result=result,
        )

    return OptimizationResponse(
        feasible=True,
        message=f"Minimum compliant power found: {result.config.power:.1f} W.",
        objective=OPTIMIZATION_OBJECTIVE,
        fixed_parameters=_fixed_parameters_for({"power"}),
        checked=checked,
        config=result.config,
        result=result,
    )


@router.post("/optimize/advanced", response_model=OptimizationResponse)
async def optimize_advanced(request: AdvancedOptimizationRequest):
    """Optimize selected installation variables against installed W/m."""
    config = request.config
    variables = request.variables
    objective = request.objective
    objective_label = _advanced_objective_label(objective)
    ldt_id, _ = require_ldt_for_config(config)

    if ldt_id.startswith("temp-"):
        exact_result = run_calculation(config, ldt_id)
        return OptimizationResponse(
            feasible=False,
            message=(
                "External LDTs are calculated exactly. Advanced optimization uses the DB reference "
                "and formulas, so select a catalog luminaire first."
            ),
            objective=objective_label,
            fixed_parameters=_fixed_parameters_for(set()),
            checked=1,
            config=config,
            result=exact_result,
        )

    return _run_advanced_search(config, variables, objective, ldt_id, objective_label)


@router.post("/optimize/advanced-batch", response_model=BatchCalculationResponse)
async def optimize_advanced_batch(request: AdvancedOptimizationRequest):
    """Optimize each selected optic and return one downloadable row per lens."""
    config = request.config
    variables = request.variables.model_copy(update={"optic_family": True})
    objective = request.objective
    objective_label = _advanced_objective_label(objective)
    items: list[BatchCalculationItem] = []

    for row, optic_family in enumerate(_optic_candidates(config, request.optic_families), start=1):
        try:
            optic_config = _with_updates(config, {"optic_family": optic_family, "ldt_id": None}, "")
            ldt_id, _ = require_ldt_for_config(optic_config)
            if ldt_id.startswith("temp-"):
                raise ValueError("External LDTs cannot be used for lens batch optimization.")
            response = _run_advanced_search(optic_config, variables, objective, ldt_id, objective_label)
            if not response.result or not response.config:
                raise ValueError(response.message)
            if not response.feasible:
                items.append(BatchCalculationItem(
                    model_id=f"{config.model_family or 'Luminaire'} {optic_family}",
                    row=row,
                    error=response.message,
                ))
                continue
            items.append(BatchCalculationItem(
                model_id=f"{config.model_family or 'Luminaire'} {optic_family}",
                row=row,
                config=response.config,
                result=response.result,
            ))
        except Exception as exc:
            items.append(BatchCalculationItem(
                model_id=f"{config.model_family or 'Luminaire'} {optic_family}",
                row=row,
                error=str(exc),
            ))

    return BatchCalculationResponse(
        filename=f"Optimized lenses - {config.manufacturer or ''} {config.model_family or ''}".strip(),
        count=len(items),
        items=items,
    )


def _float(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, str):
        value = value.strip().replace(",", ".")
        value = re.sub(r"[^0-9.+-]", "", value)
        if value in ("", "-", "+", ".", "-.", "+."):
            return default
    return float(value)


def _int(value, default=0):
    return int(round(_float(value, default)))


def _text(value, default=""):
    if value is None:
        return default
    return str(value).strip()


def _cct(value):
    match = re.search(r"\d+", _text(value))
    return int(match.group(0)) if match else 4000


def _arrangement(value):
    raw = _text(value, "Lineal").lower()
    if "bilateral" in raw:
        return "Bilateral"
    if "doble" in raw or "twin" in raw:
        return "Central Doble"
    if "isleta" in raw or "single" in raw:
        return "En Isleta"
    return "Lineal"


def _lighting_class(value):
    raw = _text(value, "M3").upper()
    match = re.search(r"\b[MP][1-6]\b", raw)
    return match.group(0) if match else "M3"


def _optic_family(value, height, road_width):
    raw = _text(value).upper()
    match = re.search(r"\bF[0-9A-Z]{2,4}\b", raw)
    if match:
        return match.group(0)
    ratio = height / road_width if road_width else 0
    if ratio <= 1:
        return "F151"
    return "F2MD"


def _config_from_row(ws, row):
    height = _float(ws.cell(row, 3).value, 9)
    road_width = _float(ws.cell(row, 6).value, 7)
    optic_family = _optic_family(ws.cell(row, 12).value, height, road_width)

    return CalculationConfig(
        road_width=road_width,
        sidewalk_left=_float(ws.cell(row, 5).value, 0),
        sidewalk_right=0,
        lanes=2,
        arrangement=_arrangement(ws.cell(row, 2).value),
        height=height,
        spacing=_float(ws.cell(row, 4).value, 30),
        arm_length=_float(ws.cell(row, 7).value, 0),
        tilt=_float(ws.cell(row, 13).value, 0),
        optic_family=optic_family,
        power=_float(ws.cell(row, 14).value, 100),
        lighting_class=_lighting_class(ws.cell(row, 8).value),
        mf=_float(ws.cell(row, 9).value, 0.85),
        pavement="R3",
        cct=_cct(ws.cell(row, 10).value),
    )


def _config_from_values(values):
    height = _float(values[2] if len(values) > 2 else None, 9)
    road_width = _float(values[5] if len(values) > 5 else None, 7)
    optic_family = _optic_family(values[11] if len(values) > 11 else None, height, road_width)

    return CalculationConfig(
        road_width=road_width,
        sidewalk_left=_float(values[4] if len(values) > 4 else None, 0),
        sidewalk_right=0,
        lanes=2,
        arrangement=_arrangement(values[1] if len(values) > 1 else None),
        height=height,
        spacing=_float(values[3] if len(values) > 3 else None, 30),
        arm_length=_float(values[6] if len(values) > 6 else None, 0),
        tilt=_float(values[12] if len(values) > 12 else None, 0),
        optic_family=optic_family,
        power=_float(values[13] if len(values) > 13 else None, 100),
        lighting_class=_lighting_class(values[7] if len(values) > 7 else None),
        mf=_float(values[8] if len(values) > 8 else None, 0.85),
        pavement="R3",
        cct=_cct(values[9] if len(values) > 9 else None),
    )


@router.post("/batch-excel", response_model=BatchCalculationResponse)
async def calculate_batch_excel(file: UploadFile = File(...)):
    """Read a DIALux-style Excel file and calculate every model row."""
    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    ws = wb.active

    rows = []
    for row_index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = tuple(values)
        model_id = _text(values[0] if values else None, f"Row {row_index}")
        if not model_id:
            continue
        rows.append((row_index, model_id, values))

    def calculate_row(row_data):
        row, model_id, values = row_data
        try:
            config = _config_from_values(values)
            ldt_id, _ = require_ldt_for_config(config)
            result = run_calculation(config, ldt_id)
            return BatchCalculationItem(model_id=model_id, row=row, config=config, result=result)
        except Exception as exc:
            return BatchCalculationItem(model_id=model_id, row=row, error=str(exc))

    with ThreadPoolExecutor(max_workers=8) as executor:
        items = list(executor.map(calculate_row, rows))

    return BatchCalculationResponse(filename=file.filename or "uploaded.xlsx", count=len(items), items=items)
