"""Batch processing: Excel parsing, dynamic LDT library, row evaluation."""
from __future__ import annotations

import io
import math
import re
from pathlib import Path
from typing import Callable, List, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .calc import Photometry, calc_luminance, calc_road, calc_SR, ME_REQ, P_REQ
from .eulumdat import parse_ldt


# ── Familia óptica desde lum_name ───────────────────────────────────────────
# Identificadores típicos en el catálogo SALVI: F151, F2MD, F2M2, F2VM, F510, etc.
# Patrón: letra F seguida de 2-4 alfanuméricos, normalmente separado por espacios o '_'.
_OPTIC_FAMILY_RE = re.compile(r"\b(F[0-9A-Z]{2,4})\b")


def extract_optic_family(photometry: "Photometry") -> str:
    """Devuelve el código de familia óptica detectado en el nombre de la luminaria.

    Ejemplos:
        'CLAP M C42 30K F2MD VDR SPUW 200W' → 'F2MD'
        'KRONOS 14G F151 VDR SPUW 35W'      → 'F151'

    Si no encuentra patrón, devuelve 'UNKNOWN'.
    """
    name = photometry.d.get("lum_name", "") or ""
    m = _OPTIC_FAMILY_RE.search(name)
    return m.group(1) if m else "UNKNOWN"


def group_photometries_by_family(
    photometries: list["Photometry"],
) -> dict[str, list["Photometry"]]:
    """Agrupa fotometrías por familia óptica, ordenadas por flujo creciente dentro
    de cada familia."""
    groups: dict[str, list[Photometry]] = {}
    for ph in photometries:
        fam = extract_optic_family(ph)
        groups.setdefault(fam, []).append(ph)
    for fam in groups:
        groups[fam].sort(key=lambda p: p.flux)
    return groups

# ── Excel column indices (0-based) ──────────────────────────────────────────
C_NAME  = 0   # A  IDENTIFICADOR MODELO
C_ARR   = 1   # B  DISPOSICIÓN
C_H     = 2   # C  ALTURA (m)
C_S     = 3   # D  INTERDISTANCIA (m)
C_WA    = 4   # E  ANCHO ACERA (m)
C_W     = 5   # F  ANCHO CALZADA (m)
C_ARM   = 6   # G  ARM LENGTH (m)
C_CLASS = 7   # H  LIGHTING CLASS
C_MF    = 8   # I  FACTOR MANTENIMIENTO
C_CCT   = 9   # J  TEMPERATURA COLOR
C_LUM   = 10  # K  LUMINARIA PROPUESTA
C_OPT   = 11  # L  ÓPTICA
C_TILT  = 12  # M  ÁNGULO INCLINACIÓN (°)
C_WPROP = 13  # N  POTENCIA PROPUESTA (W)


def parse_excel_rows(file_obj) -> list[dict]:
    """Parse an Excel workbook and return list of row dicts.

    Each dict has: name, arrangement, h, S, Wa, W, arm, class_, mf,
    cct, lum, optica, tilt, W_proposed, skip, skip_reason.
    """
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[C_NAME]:
            continue
        name = str(raw[C_NAME]).strip()
        W_val = raw[C_W]
        skip = False
        skip_reason = ""
        if not isinstance(W_val, (int, float)) or W_val != W_val or W_val == 0:
            skip = True
            skip_reason = "solo acera (W no válido)"
        rows.append(dict(
            name=name,
            arrangement=str(raw[C_ARR]).strip() if raw[C_ARR] else "Lineal",
            h=float(raw[C_H] or 6),
            S=float(raw[C_S] or 30),
            Wa=float(raw[C_WA] or 0),
            W=float(W_val) if not skip else 0.0,
            arm=float(raw[C_ARM] or 0),
            class_=str(raw[C_CLASS]).strip() if raw[C_CLASS] else "M4",
            mf=float(raw[C_MF] or 0.85),
            cct=str(raw[C_CCT] or "3000K"),
            lum=str(raw[C_LUM] or ""),
            optica=raw[C_OPT],
            tilt=float(raw[C_TILT] or 0),
            W_proposed=raw[C_WPROP],
            skip=skip,
            skip_reason=skip_reason,
        ))
    wb.close()
    return rows


def load_ldt_files(file_objects) -> list[Photometry]:
    """Load a list of file-like objects as Photometry instances.

    Each file_object must have .name (str) and .read() → bytes.
    """
    import tempfile, os
    result = []
    for fo in file_objects:
        content = fo.read()
        suffix = Path(fo.name).suffix or ".ldt"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            d = parse_ldt(tmp_path)
            result.append(Photometry(d))
        finally:
            os.unlink(tmp_path)
    return result


def build_lm_to_w(photometries: list[Photometry]) -> Callable:
    """Build a flux→watts interpolation function from a list of Photometry objects."""
    pairs = sorted(set((ph.flux, ph.power) for ph in photometries), key=lambda p: p[0])

    def lm_to_W(target_flux: float) -> tuple[float, str]:
        if not pairs:
            return target_flux / 150.0, "sin LDTs (150 lm/W asumido)"
        if target_flux <= pairs[0][0]:
            eff = pairs[0][0] / pairs[0][1]
            return target_flux / eff, f"extrap <{pairs[0][1]:.0f}W ({eff:.1f} lm/W)"
        if target_flux >= pairs[-1][0]:
            eff = pairs[-1][0] / pairs[-1][1]
            return target_flux / eff, f"extrap >{pairs[-1][1]:.0f}W ({eff:.1f} lm/W)"
        for i in range(len(pairs) - 1):
            f1, w1 = pairs[i]
            f2, w2 = pairs[i + 1]
            if f1 <= target_flux <= f2:
                w = w1 + (w2 - w1) * (target_flux - f1) / (f2 - f1)
                return w, f"interp {w1:.0f}↔{w2:.0f}W"
        return target_flux / (pairs[-1][0] / pairs[-1][1]), "?"

    return lm_to_W


def _build_family_lm_to_W(photometries: list[Photometry]) -> Callable:
    """lm_to_W interpolation function for a single optic family."""
    pairs = sorted(set((ph.flux, ph.power) for ph in photometries), key=lambda p: p[0])

    def lm_to_W(target_flux: float) -> tuple[float, str]:
        if not pairs:
            return target_flux / 150.0, "sin LDTs"
        if target_flux <= pairs[0][0]:
            eff = pairs[0][0] / pairs[0][1]
            return target_flux / eff, f"≤{pairs[0][1]:.0f}W (extrap)"
        if target_flux >= pairs[-1][0]:
            eff = pairs[-1][0] / pairs[-1][1]
            return target_flux / eff, f">{pairs[-1][1]:.0f}W (extrap)"
        for i in range(len(pairs) - 1):
            f1, w1 = pairs[i]
            f2, w2 = pairs[i + 1]
            if f1 <= target_flux <= f2:
                w = w1 + (w2 - w1) * (target_flux - f1) / (f2 - f1)
                return w, f"{w1:.0f}W↔{w2:.0f}W"
        return target_flux / (pairs[-1][0] / pairs[-1][1]), "?"

    return lm_to_W


def evaluate_row_best_optic(
    cfg: dict,
    families: dict[str, list[Photometry]],
    road: str = 'R3',
    optic_hint: Optional[str] = None,
) -> dict:
    """Evalúa la fila probando todas las familias ópticas disponibles y devuelve
    la que da mejor resultado.

    Criterio de selección:
        1. Cumplimiento total (compliant=True) tiene prioridad
        2. Entre las que cumplen, la de menor target_W
        3. Si ninguna cumple, la que más se acerque (menor W mínimo)

    `optic_hint`: código de familia preferida del Excel; sólo se usa como
    desempate si dos familias cumplen con W idéntico.

    Returns:
        dict con todos los campos de evaluate_row_dynamic + 'optic_family'
        + 'alternatives' (lista resumen del resto de familias).
    """
    if not families:
        raise ValueError("No hay familias ópticas disponibles")

    results = {}
    for fam, phs in families.items():
        if not phs:
            continue
        # Usar la fotometría de menor flujo como base distributiva
        base_ph = phs[0]
        lm_to_W_fn = _build_family_lm_to_W(phs)
        try:
            res = evaluate_row_dynamic(cfg, base_ph, lm_to_W_fn, road=road)
            res["optic_family"] = fam
            results[fam] = res
        except Exception as exc:
            results[fam] = {"error": str(exc), "optic_family": fam,
                             "compliant": False, "target_W": float("inf")}

    if not results:
        raise ValueError("Ninguna familia óptica pudo evaluarse")

    # Ordenar por (cumple, target_W, hint-match) y elegir el mejor
    def _sort_key(item):
        fam, r = item
        compliant = bool(r.get("compliant", False))
        w = r.get("target_W") or float("inf")
        if w is None or (isinstance(w, float) and math.isnan(w)):
            w = float("inf")
        hint_match = 0 if (optic_hint and fam == optic_hint) else 1
        # Queremos: cumple True primero (False=1 mayor), W ascendente, hint primero
        return (0 if compliant else 1, w, hint_match)

    sorted_results = sorted(results.items(), key=_sort_key)
    best_fam, best_res = sorted_results[0]
    best_res = dict(best_res)
    best_res["optic_family"] = best_fam

    # Resumen de alternativas (para auditoría)
    alternatives = []
    for fam, r in sorted_results[1:]:
        alternatives.append(dict(
            optic_family=fam,
            target_W=r.get("target_W"),
            compliant=r.get("compliant", False),
            error=r.get("error"),
        ))
    best_res["alternatives"] = alternatives
    return best_res


def evaluate_row_dynamic(cfg: dict, base_ph: Photometry, lm_to_W_fn: Callable,
                         road: str = 'R3') -> dict:
    """Evaluate a single row config using a dynamic lm_to_W function.

    Same logic as solver.evaluate_row but with a pluggable flux→watts mapping.
    Returns a result dict (same schema as solver.evaluate_row).

    `road` selects the CIE 144 pavement reflection table: 'R1', 'R2', 'R3' or 'R4'.
    """
    eclass = cfg["class"]
    if eclass.startswith("M"):
        rL = calc_luminance(cfg, base_ph, flux_scale=1.0, road=road)
        SR = calc_SR(cfg, base_ph, flux_scale=1.0)
        req = ME_REQ[eclass]
        scale = req["L"] / rL["Lavg"] if rL["Lavg"] > 0 else float("inf")
        TI = rL["TI"] * (scale ** 0.2)
        target_flux = base_ph.flux * scale
        W, source = lm_to_W_fn(target_flux)
        Lavg = req["L"]
        Lmin = rL["Lmin"] * scale
        Uo, Ul = rL["Uo"], rL["Ul"]
        return dict(
            mode="ME", eclass=eclass,
            Lavg=Lavg, Lmin=Lmin, Uo=Uo, Ul=Ul, TI=TI, SR=SR,
            target_flux=target_flux, target_W=W, source=source,
            ok_L=True,
            ok_Uo=Uo >= req["Uo"],
            ok_Ul=Ul >= req["Ul"],
            ok_TI=TI <= req["TI"],
            ok_SR=SR >= req["SR"],
            compliant=all([Uo >= req["Uo"], Ul >= req["Ul"], TI <= req["TI"], SR >= req["SR"]]),
            req=req,
        )
    if eclass.startswith("P"):
        rE = calc_road(cfg, base_ph, flux_scale=1.0)
        req = P_REQ.get(eclass, dict(Eavg=0, Emin=0))
        scale = req["Eavg"] / rE["Eavg"] if rE["Eavg"] > 0 else float("inf")
        target_flux = base_ph.flux * scale
        W, source = lm_to_W_fn(target_flux)
        Eavg = req["Eavg"]
        Emin = rE["Emin"] * scale
        Uo = rE["Emin"] / rE["Eavg"] if rE["Eavg"] > 0 else 0.0
        return dict(
            mode="P", eclass=eclass,
            Eavg=Eavg, Emin=Emin, Uo=Uo,
            target_flux=target_flux, target_W=W, source=source,
            ok_E=True,
            ok_Emin=Emin >= req["Emin"],
            compliant=Emin >= req["Emin"],
            req=req,
        )
    raise ValueError(f"Clase desconocida: {eclass!r}")


def run_batch(
    rows: list[dict],
    base_ph: Optional[Photometry] = None,
    lm_to_W_fn: Optional[Callable] = None,
    selected_opticas: Optional[set] = None,
    progress_cb=None,
    road: str = 'R3',
    families: Optional[dict[str, list[Photometry]]] = None,
    allowed_families: Optional[set[str]] = None,
) -> list[dict]:
    """Evaluate all rows.

    Modo recomendado: pasar `families` (dict familia → [Photometry…]).
    La app probará cada familia para cada fila y elegirá la mejor (auto-selección).
    El campo 'optica' del Excel se usa solo como pista (hint) para desempates.

    Modo legacy: pasar `base_ph` y `lm_to_W_fn` (una sola familia).
    Se mantiene por retrocompatibilidad con tests y código antiguo.

    `allowed_families`: si se pasa, sólo se consideran esas familias en el
    auto-selector (resto descartadas por configuración del usuario).
    `selected_opticas` (legacy): filtra filas del Excel por su columna 'optica'.

    Returns:
        list of dicts with keys: row_name, skip, skip_reason, error, cfg, result.
        Cuando se usa `families`, `result` incluye 'optic_family' y 'alternatives'.
    """
    if families is None and base_ph is None:
        raise ValueError("Pasa 'families' (modo recomendado) o 'base_ph'+'lm_to_W_fn' (legacy)")

    # Restringir el conjunto de familias si el usuario lo pidió
    eval_families = families
    if families is not None and allowed_families is not None:
        eval_families = {k: v for k, v in families.items() if k in allowed_families}

    total = len(rows)
    results = []
    for i, row in enumerate(rows):
        res = dict(
            row_name=row["name"],
            skip=row["skip"],
            skip_reason=row["skip_reason"],
            error=None,
            cfg=None,
            result=None,
        )
        if not row["skip"]:
            # Filtro legacy por columna 'optica' del Excel
            if selected_opticas is not None and row["optica"] not in selected_opticas:
                res["skip"] = True
                res["skip_reason"] = f"óptica {row['optica']} no seleccionada"
            else:
                cfg = dict(
                    arrangement=row["arrangement"],
                    h=row["h"],
                    S=row["S"],
                    W=row["W"],
                    Wa=row["Wa"],
                    arm=row["arm"],
                    **{"class": row["class_"]},
                    tilt=row["tilt"],
                    mf=row["mf"],
                )
                res["cfg"] = cfg
                try:
                    if eval_families:
                        # Auto-selección: probar todas las familias permitidas
                        hint = str(row.get("optica") or "").strip() or None
                        res["result"] = evaluate_row_best_optic(
                            cfg, eval_families, road=road, optic_hint=hint,
                        )
                    else:
                        # Modo legacy
                        res["result"] = evaluate_row_dynamic(
                            cfg, base_ph, lm_to_W_fn, road=road,
                        )
                except Exception as e:
                    res["error"] = str(e)
        results.append(res)
        if progress_cb:
            progress_cb(i + 1, total)
    return results


# ── Min-power batch ──────────────────────────────────────────────────────────

def _best_real_ldt(cfg, phs: list[Photometry], road: str, continuous_W: float):
    """Among real LDTs in `phs`, find the one with lowest power that meets
    all criteria. Returns (photometry, result_dict) or (None, None)."""
    best_ph = None
    best_res = None
    for ph in sorted(phs, key=lambda p: p.power):
        eclass = cfg["class"]
        if eclass.startswith("M"):
            rL = calc_luminance(cfg, ph, flux_scale=1.0, road=road)
            SR = calc_SR(cfg, ph, flux_scale=1.0)
            req = ME_REQ[eclass]
            if rL["Lavg"] >= req["L"] and rL["Uo"] >= req["Uo"] and \
               rL["Ul"] >= req["Ul"] and rL["TI"] <= req["TI"] and SR >= req["SR"]:
                best_ph = ph
        elif eclass.startswith("P"):
            rE = calc_road(cfg, ph, flux_scale=1.0)
            req = P_REQ.get(eclass, dict(Eavg=0, Emin=0))
            if rE["Eavg"] >= req["Eavg"] and rE["Emin"] >= req["Emin"]:
                best_ph = ph
        if best_ph is not None:
            break  # lowest power that works
    return best_ph


def run_batch_min_power(
    rows: list[dict],
    families: dict[str, list[Photometry]],
    allowed_families: Optional[set[str]] = None,
    progress_cb=None,
    road: str = 'R3',
) -> list[dict]:
    """Evaluate all rows in **minimum-power** mode.

    For each row, uses the continuous photometric solver to find the
    *theoretical* minimum power needed. Then maps that to the nearest
    real LDT in the chosen optic family.

    Returns
    -------
    list of dicts with keys:
        row_name, skip, skip_reason, error, cfg, result
    where ``result`` includes:
        feasible      – True if a real LDT exists that satisfies ALL criteria
        target_W      – power of the best real LDT that works (or None)
        target_flux   – flux of that LDT
        continuous_W  – theoretical optimal power (continuous)
        LDT_real      – name of the chosen real LDT
        optic_family  – chosen family
        Lavg, Uo, etc – photometric values at the LDT's native power
        failing       – list of issues if not feasible
    """
    if families is None:
        raise ValueError("'families' is required for min-power mode")

    eval_families = families
    if allowed_families is not None:
        eval_families = {k: v for k, v in families.items() if k in allowed_families}

    from .solver import find_min_power

    total = len(rows)
    results = []
    for i, row in enumerate(rows):
        res = dict(
            row_name=row["name"],
            skip=row["skip"],
            skip_reason=row["skip_reason"],
            error=None,
            cfg=None,
            result=None,
        )
        if not row["skip"]:
            cfg = dict(
                arrangement=row["arrangement"],
                h=row["h"],
                S=row["S"],
                W=row["W"],
                Wa=row["Wa"],
                arm=row["arm"],
                **{"class": row["class_"]},
                tilt=row["tilt"],
                mf=row["mf"],
            )
            res["cfg"] = cfg
            try:
                best_overall = None
                for fam, phs in eval_families.items():
                    if not phs:
                        continue
                    # 1) Continuous calculation
                    lm_to_W_fn = _build_family_lm_to_W(phs)
                    try:
                        cont = find_min_power(cfg, base_photometry=phs[0], lm_to_W_fn=lm_to_W_fn, road=road)
                    except Exception:
                        continue
                    cont_W = cont.get("target_W")
                    cont_feas = cont.get("feasible", False)
                    # Fallback: if continuous_W is None, compute Lavg-only scaling
                    if cont_W is None:
                        try:
                            from .calc import calc_luminance
                            eclass = cfg["class"]
                            if eclass.startswith("M"):
                                rL = calc_luminance(cfg, phs[0], flux_scale=1.0, road=road)
                                req = ME_REQ[eclass]
                                scale = req["L"] / rL["Lavg"] if rL["Lavg"] > 0 else float("inf")
                                if scale != float("inf"):
                                    cont_W = lm_to_W_fn(phs[0].flux * scale)[0]
                            elif eclass.startswith("P"):
                                rE = calc_road(cfg, phs[0], flux_scale=1.0)
                                req = P_REQ.get(eclass, dict(Eavg=0))
                                scale = req["Eavg"] / rE["Eavg"] if rE["Eavg"] > 0 else float("inf")
                                if scale != float("inf"):
                                    cont_W = lm_to_W_fn(phs[0].flux * scale)[0]
                        except Exception:
                            pass

                    # 2) Find the best real LDT
                    real_ph = _best_real_ldt(cfg, phs, road, cont_W)
                    if real_ph is not None:
                        # Real LDT found that works – use its native values
                        eclass = cfg["class"]
                        if eclass.startswith("M"):
                            rL = calc_luminance(cfg, real_ph, flux_scale=1.0, road=road)
                            SR = calc_SR(cfg, real_ph, flux_scale=1.0)
                            cand = dict(
                                mode="ME", eclass=eclass,
                                feasible=True,
                                target_W=real_ph.power,
                                target_flux=real_ph.flux,
                                continuous_W=cont_W,
                                Lavg=rL["Lavg"], Uo=rL["Uo"], Ul=rL["Ul"],
                                TI=rL["TI"], SR=SR,
                                failing=[], source="",
                                optic_family=fam,
                                optic_name=real_ph.d.get("lum_name", ""),
                            )
                        else:
                            rE = calc_road(cfg, real_ph, flux_scale=1.0)
                            cand = dict(
                                mode="P", eclass=eclass,
                                feasible=True,
                                target_W=real_ph.power,
                                target_flux=real_ph.flux,
                                continuous_W=cont_W,
                                Eavg=rE["Eavg"], Emin=rE["Emin"],
                                Uo=rE["Emin"] / rE["Eavg"] if rE["Eavg"] > 0 else 0,
                                failing=[], source="",
                                optic_family=fam,
                                optic_name=real_ph.d.get("lum_name", ""),
                            )
                    else:
                        # No real LDT works – show continuous result + failing reasons
                        cand = dict(cont)
                        cand["feasible"] = False  # no real LDT found
                        cand["optic_family"] = fam
                        cand["target_W"] = None
                        cand["target_flux"] = None
                        cand["continuous_W"] = cont_W
                        cand["optic_name"] = ""
                        # Add info about what real LDT would be needed
                        if cont_W is not None:
                            cand["failing"] = cont.get("failing", [])
                            if not cand["failing"]:
                                cand["failing"] = [f"Se necesitan {cont_W:.0f}W pero el máximo LDT es {max(p.power for p in phs):.0f}W"]

                    if best_overall is None:
                        best_overall = cand
                    else:
                        # Prefer feasible, then lower continuous_W
                        if cand["feasible"] and not best_overall["feasible"]:
                            best_overall = cand
                        elif cand["feasible"] == best_overall["feasible"]:
                            cw = cand.get("continuous_W") or float("inf")
                            bw = best_overall.get("continuous_W") or float("inf")
                            if cw < bw:
                                best_overall = cand

                if best_overall is None:
                    raise ValueError("Ninguna familia pudo evaluarse")
                res["result"] = best_overall
            except Exception as e:
                res["error"] = str(e)
        results.append(res)
        if progress_cb:
            progress_cb(i + 1, total)
    return results


# ── Excel output ─────────────────────────────────────────────────────────────

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED   = PatternFill("solid", fgColor="FFC7CE")
_GRAY  = PatternFill("solid", fgColor="EBEBEB")
_BOLD  = Font(bold=True)
_CENTER = Alignment(horizontal="center")

_HDR_FILL = PatternFill("solid", fgColor="1F497D")
_HDR_FONT = Font(bold=True, color="FFFFFF")


def write_results_xlsx(rows: list[dict], batch_results: list[dict]) -> bytes:
    """Write a summary Excel with input data + calculated results.

    Returns bytes of the .xlsx file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = [
        "MODELO", "DISPOSICIÓN", "h (m)", "S (m)", "Wa (m)", "W (m)",
        "arm (m)", "CLASE", "MF", "ÓPTICA Excel (hint)", "TILT (°)",
        "ÓPTICA ELEGIDA",
        "RESULTADO",
        "W mínima (W)", "Φ (lm)",
        "Lavg/Eavg", "Uo", "Ul/–", "TI (%) /–", "SR/–",
        "ok_L/E", "ok_Uo", "ok_Ul", "ok_TI", "ok_SR",
        "LDT SELECCIONADO",
        "NOTAS",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTER

    for r_idx, (row, br) in enumerate(zip(rows, batch_results), 2):
        res = br.get("result")
        err = br.get("error")
        skip = br.get("skip")
        skip_reason = br.get("skip_reason", "")

        def _cell(col, val, fill=None):
            c = ws.cell(row=r_idx, column=col, value=val)
            if fill:
                c.fill = fill
            return c

        _cell(1, row["name"])
        _cell(2, row["arrangement"])
        _cell(3, row["h"])
        _cell(4, row["S"])
        _cell(5, row["Wa"])
        _cell(6, row["W"] if not skip else "–")
        _cell(7, row["arm"])
        _cell(8, row["class_"])
        _cell(9, row["mf"])
        _cell(10, row["optica"])
        _cell(11, row["tilt"])

        if skip:
            _cell(12, "–")
            _cell(13, "OMITIDO", _GRAY)
            _cell(27, skip_reason)
        elif err:
            _cell(12, "–")
            _cell(13, "ERROR", _RED)
            _cell(27, err)
        elif res:
            is_min_power = "feasible" in res
            if is_min_power:
                feasible = res.get("feasible", False)
                cw = res.get("continuous_W")
                fill = _GREEN if feasible else _RED
                _cell(12, res.get("optic_family", "?"))
                if feasible:
                    _cell(13, "✓ FACTIBLE", fill)
                    _cell(14, round(res["target_W"], 1))
                    _cell(15, round(res["target_flux"], 0))
                else:
                    _cell(13, "✗ NO FACTIBLE", fill)
                    _cell(14, round(cw, 1) if cw else "–")
                    _cell(15, "–")
                if res.get("Lavg") is not None:  # ME mode
                    _cell(16, round(res["Lavg"], 3) if res["Lavg"] is not None else "–")
                    _cell(17, round(res["Uo"], 3) if res["Uo"] is not None else "–")
                    _cell(18, round(res["Ul"], 3) if res["Ul"] is not None else "–")
                    _cell(19, round(res["TI"], 1) if res["TI"] is not None else "–")
                    _cell(20, round(res["SR"], 3) if res["SR"] is not None else "–")
                    for col, key in [(21, "ok_L"), (22, "ok_Uo"), (23, "ok_Ul"), (24, "ok_TI"), (25, "ok_SR")]:
                        _cell(col, "–")
                else:  # P mode
                    _cell(16, round(res["Eavg"], 3) if res.get("Eavg") else "–")
                    _cell(17, round(res["Uo"], 3) if res.get("Uo") else "–")
                    _cell(18, "–"); _cell(19, "–"); _cell(20, "–")
                    _cell(21, "–"); _cell(22, "–"); _cell(23, "–"); _cell(24, "–"); _cell(25, "–")
                lname = res.get("optic_name", "")
                _cell(26, lname)
                if feasible:
                    note = f"Potencia necesaria: {cw:.0f}W · LDT real: {res['target_W']:.0f}W ({lname})" if cw else f"LDT real: {res['target_W']:.0f}W ({lname})"
                    _cell(27, note)
                else:
                    failing = "; ".join(res.get("failing", []))
                    note = f"{failing} | Potencia necesaria: {cw:.0f}W" if cw else failing
                    _cell(27, note)
            else:
                compliant = res.get("compliant", False)
                fill = _GREEN if compliant else _RED
                _cell(12, res.get("optic_family", "?"))
                _cell(13, "✓ CUMPLE" if compliant else "✗ NO CUMPLE", fill)
                _cell(14, round(res["target_W"], 1))
                _cell(15, round(res["target_flux"], 0))
                if res["mode"] == "ME":
                    _cell(16, round(res["Lavg"], 3))
                    _cell(17, round(res["Uo"], 3))
                    _cell(18, round(res["Ul"], 3))
                    _cell(19, round(res["TI"], 1))
                    _cell(20, round(res["SR"], 3))
                    for col, key in [(21, "ok_L"), (22, "ok_Uo"), (23, "ok_Ul"), (24, "ok_TI"), (25, "ok_SR")]:
                        ok = res.get(key, False)
                        _cell(col, "✓" if ok else "✗", _GREEN if ok else _RED)
                else:
                    _cell(16, round(res["Eavg"], 3))
                    _cell(17, round(res["Uo"], 3))
                    _cell(18, "–"); _cell(19, "–"); _cell(20, "–")
                    for col, key in [(21, "ok_E"), (23, "ok_Emin")]:
                        ok = res.get(key, False)
                        _cell(col, "✓" if ok else "✗", _GREEN if ok else _RED)
                    _cell(22, "–"); _cell(24, "–"); _cell(25, "–")
                _cell(26, res.get("source", ""))

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
