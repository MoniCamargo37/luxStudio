from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..schemas.models import CalculationResult
from .i18n import translator


def _headers(t):
    return [
        t("excel.model_id"),
        t("excel.arrangement"),
        t("excel.height"),
        t("excel.spacing"),
        t("excel.sidewalk_width"),
        t("excel.road_width"),
        t("excel.arm_length"),
        t("excel.lighting_class"),
        t("excel.maintenance"),
        t("excel.cct"),
        "CRI",
        t("excel.proposed_luminaire"),
        t("excel.optic"),
        t("excel.tilt"),
        t("excel.proposed_power"),
        "Lm (cd/m2)\nEm (lux)",
        "UNIFORMIDAD\nUo\nEmin (Lux)",
        "UI",
        "TI",
        "SR",
        t("excel.project_efficiency"),
        "int/h",
        "h/a",
        None,
        "P_calc\n(W)",
        "Phi_calc\n(lm)",
        "CRI_calc",
        "LDT base\n/interp",
        "Lm/Em\ncalc",
        "Uo\ncalc",
        "UI/Ul\ncalc",
        "TI\ncalc",
        "SR\ncalc",
        t("excel.complies"),
        t("excel.notes"),
    ]


def _metric(result: CalculationResult, key: str):
    value = getattr(result, key, None)
    return round(value, 3) if value is not None else None


def _notes(result: CalculationResult) -> str:
    failed = [c.name for c in result.criteria if not c.passed]
    return ", ".join(failed) if failed else "OK"


def _spacing_note(lighting_class: str, spacing: float, road_width: float, t) -> str:
    if road_width <= 0:
        return ""
    ratio = spacing / road_width
    if lighting_class in ("M1", "M2"):
        return "OK m1-m2" if ratio < 4 else t("excel.excessive_spacing")
    if lighting_class in ("M3", "M4", "M5"):
        return "OK m3-m4-m5" if ratio < 4.5 else t("excel.excessive_spacing")
    return ""


def _optic_note(height: float, road_width: float, t) -> str:
    if road_width <= 0:
        return ""
    ratio = height / road_width
    if ratio <= 1:
        return "f151"
    if ratio > 1.5:
        return t("excel.watch")
    return "f2md"


def _sidewalk_value(left: float, right: float):
    return left if abs(left - right) < 0.001 else f"L {left:.1f} / R {right:.1f}"


def _row_values(result: CalculationResult, t):
    cfg = result.config
    lum = result.luminaire
    main_metric = result.Lavg if result.mode == "ME" else result.Eavg
    uniformity = result.Uo if result.mode == "ME" else result.Emin

    return [
        "MODELO 1",
        cfg.arrangement,
        cfg.height,
        cfg.spacing,
        _sidewalk_value(cfg.sidewalk_left, cfg.sidewalk_right),
        cfg.road_width,
        cfg.arm_length,
        cfg.lighting_class,
        cfg.mf,
        f"{cfg.cct}K",
        cfg.cri,
        lum.luminaire_name,
        lum.optic_family,
        cfg.tilt,
        cfg.power,
        _metric(result, "Lavg") if result.mode == "ME" else _metric(result, "Eavg"),
        _metric(result, "Uo") if result.mode == "ME" else _metric(result, "Emin"),
        _metric(result, "Ul"),
        _metric(result, "TI"),
        _metric(result, "SR"),
        None,
        _spacing_note(cfg.lighting_class, cfg.spacing, cfg.road_width, t),
        _optic_note(cfg.height, cfg.road_width, t),
        None,
        lum.power,
        round(lum.flux, 0),
        lum.cri,
        lum.filename,
        round(main_metric, 3) if main_metric is not None else None,
        round(uniformity, 3) if uniformity is not None else None,
        _metric(result, "Ul"),
        _metric(result, "TI"),
        _metric(result, "SR"),
        t("status.pass_short") if result.compliant else t("status.fail_short"),
        _notes(result),
    ]


def generate_excel(result: CalculationResult) -> bytes:
    t = translator(result.config.language)
    headers = _headers(t)
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    ws.append(headers)
    ws.append(_row_values(result, t))
    ws["U2"] = '=IFERROR(D2*F2*P2*15/O2,"")'

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    fills = {
        "input": PatternFill("solid", fgColor="C6E0B4"),
        "selection": PatternFill("solid", fgColor="FFD966"),
        "spacer": PatternFill("solid", fgColor="FFFFFF"),
        "calc": PatternFill("solid", fgColor="2F5597"),
    }
    thin = Side(style="thin", color="A6A6A6")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        if col <= 8:
            cell.fill = fills["input"]
            font_color = "000000"
        elif col <= 23:
            cell.fill = fills["selection"]
            font_color = "000000"
        elif col == 24:
            cell.fill = fills["spacer"]
            font_color = "000000"
        else:
            cell.fill = fills["calc"]
            font_color = "FFFFFF"
        cell.font = Font(name="Calibri", size=8, bold=True, color=font_color)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border

        value_cell = ws.cell(2, col)
        value_cell.font = Font(name="Calibri", size=9, bold=col in (34, 35))
        value_cell.alignment = Alignment(wrap_text=True, vertical="center")
        value_cell.border = border
        if col == 34:
            value_cell.fill = PatternFill("solid", fgColor="E2F0D9" if result.compliant else "F4CCCC")

    number_formats = {
        "C": "0.00",
        "D": "0.0",
        "E": "0.0",
        "F": "0.0",
        "G": "0.0",
        "I": "0.00",
        "K": "0",
        "N": "0",
        "O": "0",
        "P": "0.000",
        "Q": "0.000",
        "R": "0.000",
        "S": "0.0",
        "T": "0.000",
        "U": "0.00",
        "Y": "0.0",
        "Z": "#,##0",
        "AA": "0",
        "AC": "0.000",
        "AD": "0.000",
        "AE": "0.000",
        "AF": "0.0",
        "AG": "0.000",
    }
    for col, number_format in number_formats.items():
        ws[f"{col}2"].number_format = number_format

    widths = {
        "A": 16, "B": 14, "C": 10, "D": 15, "E": 15, "F": 15, "G": 13,
        "H": 14, "I": 18, "J": 18, "K": 8, "L": 28, "M": 14, "N": 16,
        "O": 16, "P": 14, "Q": 16, "R": 10, "S": 8, "T": 8, "U": 16,
        "V": 22, "W": 12, "X": 5, "Y": 11, "Z": 13, "AA": 10, "AB": 30,
        "AC": 13, "AD": 12, "AE": 12, "AF": 10, "AG": 10, "AH": 10, "AI": 30,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 38
    ws.auto_filter.ref = "A1:AI2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
